"""
ReportService — generates Excel/PDF reports for recruitment analytics.

Supports report types:
  - funnel:  Recruitment funnel analysis
  - trend:   Time-series trend data
  - position: Per-position performance
  - channel:  Channel ROI analysis
  - full:     Comprehensive report combining all above

Uses openpyxl for Excel generation (graceful fallback if not installed)
and WeasyPrint for PDF generation (graceful fallback if not installed).
"""

from __future__ import annotations

import io
import logging
import os
import tempfile
import uuid
from datetime import datetime, timezone
from typing import Any

from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.candidate import Candidate
from app.models.interview import Interview
from app.models.offer import Offer
from app.models.position import Position

logger = logging.getLogger("hiremind.services.report_service")

# In-memory report storage (replace with file storage / MinIO in production)
_report_store: dict[str, dict[str, Any]] = {}


class ReportService:
    """Async service for generating recruitment analytics reports."""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ════════════════════════════════════════════════════════════
    # Excel Report Generation
    # ════════════════════════════════════════════════════════════

    async def generate_excel_report(
        self,
        report_type: str,
        filters: dict | None,
        tenant_id: str,
    ) -> dict[str, Any]:
        """Generate an Excel report.

        Args:
            report_type: funnel / trend / position / channel / full.
            filters: Optional filters (date_from, date_to, position_id).
            tenant_id: Tenant scope.

        Returns:
            Dict with report_id, file_path or data, and metadata.
        """
        filters = filters or {}

        # Fetch data based on report type
        data = await self._gather_report_data(report_type, filters, tenant_id)

        # Try openpyxl, fallback to CSV-like mock
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            if report_type == "full":
                sheets = ["漏斗分析", "趋势数据", "岗位绩效", "渠道ROI"]
                data_map = {
                    "漏斗分析": data.get("funnel", {}),
                    "趋势数据": data.get("trend", {}),
                    "岗位绩效": data.get("position", {}),
                    "渠道ROI": data.get("channel", {}),
                }
                for i, sheet_name in enumerate(sheets):
                    if i == 0:
                        ws = wb.active
                        ws.title = sheet_name
                    else:
                        ws = wb.create_sheet(sheet_name)
                    self._write_sheet(ws, sheet_name, data_map.get(sheet_name, {}))
            else:
                sheet_name_map = {
                    "funnel": "漏斗分析",
                    "trend": "趋势数据",
                    "position": "岗位绩效",
                    "channel": "渠道ROI",
                }
                ws = wb.active
                ws.title = sheet_name_map.get(report_type, "报告")
                self._write_sheet(ws, ws.title, data)

            # Save to temp file
            report_id = str(uuid.uuid4())
            file_path = os.path.join(
                tempfile.gettempdir(), f"hiremind_report_{report_id}.xlsx"
            )
            wb.save(file_path)

            _report_store[report_id] = {
                "report_id": report_id,
                "report_type": report_type,
                "format": "excel",
                "file_path": file_path,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": tenant_id,
            }

            return {
                "report_id": report_id,
                "format": "excel",
                "file_path": file_path,
                "status": "success",
            }

        except ImportError:
            logger.warning("openpyxl not installed, generating mock report")

            report_id = str(uuid.uuid4())
            _report_store[report_id] = {
                "report_id": report_id,
                "report_type": report_type,
                "format": "excel",
                "data": data,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": tenant_id,
                "note": "openpyxl not installed; data returned as JSON mock",
            }

            return {
                "report_id": report_id,
                "format": "excel",
                "data": data,
                "status": "mock",
                "note": "openpyxl not installed; data returned as JSON",
            }

    def _write_sheet(self, ws, title: str, data: dict) -> None:
        """Write data into an openpyxl worksheet."""
        from openpyxl.styles import Font, PatternFill

        # Title row
        ws.append([title])
        ws["A1"].font = Font(bold=True, size=14)

        row = 3

        if "stages" in data:
            # Funnel data
            ws.append(["阶段", "数量", "转化率", "累计转化率"])
            for cell in ws[row]:
                if cell.value:
                    cell.font = Font(bold=True)
            row += 1
            for stage in data.get("stages", []):
                ws.append([
                    stage.get("stage", ""),
                    stage.get("count", 0),
                    stage.get("conversion_rate", ""),
                    stage.get("cumulative_rate", ""),
                ])
                row += 1

        elif "data_points" in data:
            # Trend data
            ws.append(["日期", "简历", "匹配", "面试", "Offer", "入职"])
            for cell in ws[row]:
                if cell.value:
                    cell.font = Font(bold=True)
            row += 1
            for dp in data.get("data_points", []):
                ws.append([
                    dp.get("date", ""),
                    dp.get("resumes", 0),
                    dp.get("matches", 0),
                    dp.get("interviews", 0),
                    dp.get("offers", 0),
                    dp.get("hires", 0),
                ])
                row += 1

        elif "positions" in data:
            # Position data
            ws.append(["岗位", "状态", "部门", "在招人数", "已到岗", "候选人", "面试", "Offer", "入职", "到岗率"])
            for cell in ws[row]:
                if cell.value:
                    cell.font = Font(bold=True)
            row += 1
            for pos in data.get("positions", []):
                ws.append([
                    pos.get("title", ""),
                    pos.get("status", ""),
                    pos.get("department", ""),
                    pos.get("headcount", 0),
                    pos.get("filled_count", 0),
                    pos.get("candidates", 0),
                    pos.get("interviews", 0),
                    pos.get("offers", 0),
                    pos.get("hired", 0),
                    pos.get("fill_rate", 0),
                ])
                row += 1

        elif "channels" in data:
            # Channel ROI data
            ws.append(["渠道", "总候选人", "已入职", "转化率", "单聘成本", "ROI分数"])
            for cell in ws[row]:
                if cell.value:
                    cell.font = Font(bold=True)
            row += 1
            for ch in data.get("channels", []):
                ws.append([
                    ch.get("source", ch.get("channel", "")),
                    ch.get("total", ch.get("resumes", 0)),
                    ch.get("hired", 0),
                    ch.get("conversion_rate", 0),
                    ch.get("cost_per_hire", ""),
                    ch.get("roi_score", 0),
                ])
                row += 1

        else:
            # Generic data
            ws.append(["数据"])
            ws.append([str(data)])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # ════════════════════════════════════════════════════════════
    # PDF Report Generation
    # ════════════════════════════════════════════════════════════

    async def generate_pdf_report(
        self,
        report_type: str,
        filters: dict | None,
        tenant_id: str,
    ) -> dict[str, Any]:
        """Generate a PDF report.

        Uses WeasyPrint if available, otherwise returns HTML mock.

        Args:
            report_type: funnel / trend / position / channel / full.
            filters: Optional filters.
            tenant_id: Tenant scope.

        Returns:
            Dict with report_id, file_path or html, and metadata.
        """
        filters = filters or {}
        data = await self._gather_report_data(report_type, filters, tenant_id)

        # Generate HTML content
        html_content = self._generate_html_report(report_type, data, tenant_id)

        try:
            from weasyprint import HTML

            report_id = str(uuid.uuid4())
            file_path = os.path.join(
                tempfile.gettempdir(), f"hiremind_report_{report_id}.pdf"
            )
            HTML(string=html_content).write_pdf(file_path)

            _report_store[report_id] = {
                "report_id": report_id,
                "report_type": report_type,
                "format": "pdf",
                "file_path": file_path,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": tenant_id,
            }

            return {
                "report_id": report_id,
                "format": "pdf",
                "file_path": file_path,
                "status": "success",
            }

        except ImportError:
            logger.warning("weasyprint not installed, returning HTML mock")

            report_id = str(uuid.uuid4())
            _report_store[report_id] = {
                "report_id": report_id,
                "report_type": report_type,
                "format": "pdf",
                "html": html_content,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": tenant_id,
                "note": "weasyprint not installed; HTML returned as mock",
            }

            return {
                "report_id": report_id,
                "format": "pdf",
                "html": html_content,
                "status": "mock",
                "note": "weasyprint not installed; HTML returned as mock",
            }

    def _generate_html_report(
        self, report_type: str, data: dict, tenant_id: str,
    ) -> str:
        """Generate HTML content for a report."""
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
        title_map = {
            "funnel": "招聘漏斗分析报告",
            "trend": "招聘趋势分析报告",
            "position": "岗位绩效分析报告",
            "channel": "渠道ROI分析报告",
            "full": "招聘综合分析报告",
        }
        title = title_map.get(report_type, "招聘分析报告")

        html_parts = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'>",
            f"<title>{title}</title>",
            "<style>",
            "body { font-family: 'PingFang SC', 'Microsoft YaHei', sans-serif; margin: 40px; }",
            "h1 { color: #1890ff; border-bottom: 2px solid #1890ff; padding-bottom: 10px; }",
            "h2 { color: #333; margin-top: 30px; }",
            "table { border-collapse: collapse; width: 100%; margin: 15px 0; }",
            "th { background-color: #f0f5ff; padding: 10px; text-align: left; border: 1px solid #ddd; }",
            "td { padding: 8px; border: 1px solid #ddd; }",
            "tr:nth-child(even) { background-color: #fafafa; }",
            ".meta { color: #666; font-size: 0.9em; }",
            "</style></head><body>",
            f"<h1>{title}</h1>",
            f"<p class='meta'>租户: {tenant_id} | 生成时间: {now}</p>",
        ]

        sections = [(report_type, data)] if report_type != "full" else [
            ("funnel", data.get("funnel", {})),
            ("trend", data.get("trend", {})),
            ("position", data.get("position", {})),
            ("channel", data.get("channel", {})),
        ]

        for sec_type, sec_data in sections:
            html_parts.append(self._render_section_html(sec_type, sec_data))

        html_parts.append("</body></html>")
        return "\n".join(html_parts)

    def _render_section_html(self, section_type: str, data: dict) -> str:
        """Render a single report section as HTML."""
        parts = []

        if section_type == "funnel" and "stages" in data:
            parts.append("<h2>漏斗分析</h2><table>")
            parts.append("<tr><th>阶段</th><th>数量</th><th>转化率</th><th>累计转化率</th></tr>")
            for s in data.get("stages", []):
                parts.append(
                    f"<tr><td>{s.get('stage','')}</td><td>{s.get('count',0)}</td>"
                    f"<td>{s.get('conversion_rate','-')}</td><td>{s.get('cumulative_rate','-')}</td></tr>"
                )
            parts.append("</table>")

        elif section_type == "trend" and "data_points" in data:
            parts.append("<h2>趋势数据</h2><table>")
            parts.append("<tr><th>日期</th><th>简历</th><th>匹配</th><th>面试</th><th>Offer</th><th>入职</th></tr>")
            for dp in data.get("data_points", []):
                parts.append(
                    f"<tr><td>{dp.get('date','')}</td><td>{dp.get('resumes',0)}</td>"
                    f"<td>{dp.get('matches',0)}</td><td>{dp.get('interviews',0)}</td>"
                    f"<td>{dp.get('offers',0)}</td><td>{dp.get('hires',0)}</td></tr>"
                )
            parts.append("</table>")

        elif section_type == "position" and "positions" in data:
            parts.append("<h2>岗位绩效</h2><table>")
            parts.append("<tr><th>岗位</th><th>状态</th><th>在招</th><th>已到岗</th><th>候选人</th><th>到岗率</th></tr>")
            for p in data.get("positions", []):
                parts.append(
                    f"<tr><td>{p.get('title','')}</td><td>{p.get('status','')}</td>"
                    f"<td>{p.get('headcount',0)}</td><td>{p.get('filled_count',0)}</td>"
                    f"<td>{p.get('candidates',0)}</td><td>{p.get('fill_rate',0)}%</td></tr>"
                )
            parts.append("</table>")

        elif section_type == "channel" and "channels" in data:
            parts.append("<h2>渠道ROI</h2><table>")
            parts.append("<tr><th>渠道</th><th>候选人</th><th>已入职</th><th>ROI分数</th></tr>")
            for c in data.get("channels", []):
                parts.append(
                    f"<tr><td>{c.get('source', c.get('channel',''))}</td>"
                    f"<td>{c.get('total', c.get('resumes',0))}</td>"
                    f"<td>{c.get('hired',0)}</td><td>{c.get('roi_score',0)}</td></tr>"
                )
            parts.append("</table>")

        if not parts:
            parts.append(f"<p>暂无{section_type}数据</p>")

        return "\n".join(parts)

    # ════════════════════════════════════════════════════════════
    # Data Gathering
    # ════════════════════════════════════════════════════════════

    async def _gather_report_data(
        self,
        report_type: str,
        filters: dict,
        tenant_id: str,
    ) -> dict:
        """Gather data for report generation based on type."""
        from app.services.analytics_service import AnalyticsService

        svc = AnalyticsService(self.db)
        date_from = filters.get("date_from")
        date_to = filters.get("date_to")
        position_id = filters.get("position_id")

        if report_type == "funnel":
            return await svc.get_pipeline_funnel(tenant_id, date_from, date_to)

        elif report_type == "trend":
            from app.services.dashboard_service import get_trends
            from datetime import date as date_type
            sd = str(date_from) if date_from else None
            ed = str(date_to) if date_to else None
            trend_resp = await get_trends(tenant_id, self.db, "daily", sd, ed)
            return trend_resp.model_dump()

        elif report_type == "position":
            import uuid as _uuid
            pid = _uuid.UUID(position_id) if position_id else None
            results = await svc.get_position_metrics(tenant_id, pid)
            return {"positions": results}

        elif report_type == "channel":
            return await svc.get_channel_roi(tenant_id)

        elif report_type == "full":
            return {
                "funnel": await svc.get_pipeline_funnel(tenant_id, date_from, date_to),
                "trend": (await get_trends(
                    tenant_id, self.db, "daily",
                    str(date_from) if date_from else None,
                    str(date_to) if date_to else None,
                )).model_dump() if True else {},
                "position": {"positions": await svc.get_position_metrics(tenant_id)},
                "channel": await svc.get_channel_roi(tenant_id),
            }

        return {}

    # ════════════════════════════════════════════════════════════
    # Single Position Report
    # ════════════════════════════════════════════════════════════

    async def generate_position_report(
        self,
        position_id: str,
        tenant_id: str,
    ) -> dict[str, Any]:
        """Generate a detailed single-position report.

        Args:
            position_id: UUID of the position.
            tenant_id: Tenant scope.

        Returns:
            Dict with comprehensive position analytics.
        """
        from uuid import UUID
        from app.services.analytics_service import AnalyticsService

        svc = AnalyticsService(self.db)
        analytics = await svc.get_position_analytics(
            position_id=UUID(position_id),
            tenant_id=tenant_id,
        )

        if "error" in analytics:
            return analytics

        # Add additional report metadata
        analytics["report_type"] = "position_detail"
        analytics["generated_at"] = datetime.now(timezone.utc).isoformat()

        return analytics

    # ════════════════════════════════════════════════════════════
    # Funnel Chart Export
    # ════════════════════════════════════════════════════════════

    async def export_funnel_chart(
        self,
        position_id: str | None,
        tenant_id: str,
    ) -> dict[str, Any]:
        """Export funnel visualization data.

        Returns structured data suitable for chart rendering
        (frontend can use ECharts / Chart.js).

        Args:
            position_id: Optional position UUID.
            tenant_id: Tenant scope.

        Returns:
            Dict with chart configuration and data.
        """
        from uuid import UUID
        from app.services.dashboard_service import get_funnel

        funnel_resp = await get_funnel(
            tenant_id, self.db,
            UUID(position_id) if position_id else None,
            "all",
        )

        # Transform to chart-friendly format
        stages = funnel_resp.stages
        chart_data = {
            "chart_type": "funnel",
            "title": "招聘漏斗",
            "data": [
                {
                    "name": s.label,
                    "value": s.count,
                    "conversion_rate": s.conversion_rate,
                    "cumulative_rate": s.cumulative_rate,
                }
                for s in stages
            ],
            "position_id": str(funnel_resp.position_id) if funnel_resp.position_id else None,
            "generated_at": funnel_resp.generated_at.isoformat(),
        }

        return chart_data

    # ════════════════════════════════════════════════════════════
    # Report Download
    # ════════════════════════════════════════════════════════════

    async def get_report(self, report_id: str) -> dict[str, Any] | None:
        """Retrieve a stored report by ID.

        Args:
            report_id: Report UUID.

        Returns:
            Report metadata dict or None if not found.
        """
        return _report_store.get(report_id)
